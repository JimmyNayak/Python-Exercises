import openpyxl as xl
from openpyxl.chart import BarChart, Reference

work_book = xl.load_workbook('material/transactions.xlsx')
sheet = work_book['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1, 1)

# print(cell.value)
#
# print(sheet.max_row)

sum_value = 0
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    sum_value += cell.value
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
    # print(cell.value)
# print(sum_value)


chart_values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

chart = BarChart()
chart.add_data(chart_values)

sheet.add_chart = (chart, 'e2')

work_book.save('material/transactions2.xlsx')
